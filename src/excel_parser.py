"""
Excel parser for GA4 event tracking reports — v2.

Block discovery per §1.2 of REPORT_FORMAT.md:
  - Scan r <= 30, c <= 12 for cells whose value is "Click ID" or "Event name"
  - Each such cell anchors a block of that type
  - Parse BOTH block types: click_id and event_name

Output structure (per project):
{
    "TVS Altura": {
        "pages": {
            "lp": {
                "title": "TVS Emerald Altura LP",
                "total_clicks": 193033,   # from Total row
                "total_users": 18452,
                "items": [{"click_id": str, "clicks": int, "users": int}, ...],
                "scroll": {
                    "page_view_users": 23202,
                    "buckets": [(0.40, 5770), (0.60, 4488), ...],
                } | None,
                "mode": "full" | "clicks_only",
            },
            "project": { ... } | None,
        },
        "flags": ["..."]
    }
}
"""

import re
import logging
import openpyxl

log = logging.getLogger('excel_parser')

# Regex for scroll depth bucket rows (note the intentional "Dept" typo in GA4 exports)
_SCROLL_BUCKET_RE = re.compile(r'^(\d{1,3})%\s*Scroll\s*Dept', re.IGNORECASE)

# Sheets to skip
_SKIP_SHEETS = frozenset({'index', 'summary', 'cover', 'readme'})

# GA4 system event names that are not CTAs
_GA4_SYSTEM_EVENTS = frozenset({
    'page_view', 'session_start', 'first_visit', 'user_engagement',
    'scroll', 'click', 'file_download', 'form_submit', 'video_start',
    'video_complete', 'video_progress', 'view_search_results',
})


def _cell_str(cell) -> str:
    """Return stripped string value of a cell, or '' if empty/None."""
    if cell is None or cell.value is None:
        return ''
    return str(cell.value).strip()


def _to_int(value) -> int:
    try:
        if value is None:
            return 0
        return int(float(str(value)))
    except (ValueError, TypeError):
        return 0


def _is_total_row(key_val: str) -> bool:
    return key_val.lower() in {'total', '(not set)'}


def _is_system_event(key_val: str) -> bool:
    return key_val.lower() in _GA4_SYSTEM_EVENTS


def _get_title_above(ws, header_row: int, key_col: int) -> str:
    """
    Return the first non-empty cell value found above the header row
    in the same column (checks r-1 then r-2).
    """
    for r in (header_row - 1, header_row - 2):
        if r < 1:
            continue
        v = _cell_str(ws.cell(row=r, column=key_col))
        if v and v.lower() not in {'none', 'nan'}:
            return v
    return ''


def _parse_click_id_block(ws, header_row: int, key_col: int) -> dict:
    """
    Parse one Click ID block.

    Returns:
        {
            'title': str,
            'total_clicks': int,
            'total_users': int,
            'items': [{'click_id': str, 'clicks': int, 'users': int}, ...],
        }
    """
    title = _get_title_above(ws, header_row, key_col)
    count_col = key_col + 1
    users_col = key_col + 2

    total_clicks = 0
    total_users = 0
    items = []

    row = header_row + 1
    while True:
        cell_val = _cell_str(ws.cell(row=row, column=key_col))
        if not cell_val:
            # Allow a few blank rows before giving up (merged cells etc.)
            # If two consecutive rows are empty, stop.
            if not _cell_str(ws.cell(row=row + 1, column=key_col)):
                break
            row += 1
            continue

        clicks = _to_int(ws.cell(row=row, column=count_col).value)
        users = _to_int(ws.cell(row=row, column=users_col).value)

        if _is_total_row(cell_val):
            # First data row is the Total metadata row
            total_clicks = clicks
            total_users = users
        else:
            # Exclude GA4 system events from CTA list
            if not _is_system_event(cell_val):
                items.append({
                    'click_id': cell_val,
                    'clicks': clicks,
                    'users': users,
                })

        row += 1

    return {
        'title': title,
        'total_clicks': total_clicks,
        'total_users': total_users,
        'items': items,
    }


def _parse_event_name_block(ws, header_row: int, key_col: int) -> dict:
    """
    Parse one Event name block.

    Returns:
        {
            'title': str,
            'page_view_users': int,
            'buckets': [(depth_float, users_int), ...],   # sorted asc by depth
        }
    """
    title = _get_title_above(ws, header_row, key_col)
    count_col = key_col + 1
    users_col = key_col + 2

    page_view_users = 0
    buckets = []

    row = header_row + 1
    while True:
        cell_val = _cell_str(ws.cell(row=row, column=key_col))
        if not cell_val:
            if not _cell_str(ws.cell(row=row + 1, column=key_col)):
                break
            row += 1
            continue

        users = _to_int(ws.cell(row=row, column=users_col).value)

        if cell_val.lower() == 'page_view':
            page_view_users = users
        else:
            m = _SCROLL_BUCKET_RE.match(cell_val)
            if m:
                depth = int(m.group(1)) / 100.0
                buckets.append((depth, users))
            # bare "scroll" and other system events are ignored here

        row += 1

    buckets.sort(key=lambda x: x[0])

    return {
        'title': title,
        'page_view_users': page_view_users,
        'buckets': buckets,
    }


def _discover_blocks(ws) -> list[dict]:
    """
    Scan the worksheet for Click ID / Event name header cells.

    Returns a list of dicts:
        {'kind': 'click_id'|'event_name', 'header_row': int, 'key_col': int}
    """
    blocks = []
    max_row = min(ws.max_row or 30, 30)
    max_col = min(ws.max_column or 12, 12)

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = _cell_str(ws.cell(row=r, column=c))
            if v == 'Click ID':
                blocks.append({'kind': 'click_id', 'header_row': r, 'key_col': c})
            elif v == 'Event name':
                blocks.append({'kind': 'event_name', 'header_row': r, 'key_col': c})

    return blocks


def _assign_page_type(title: str, col: int) -> str | None:
    """
    Infer whether a block belongs to the LP or Project page.

    Heuristics (in order):
      1. Title contains 'lp' → 'lp'
      2. Title contains 'project' → 'project'
      3. Column B (col 2) → 'lp', Column F (col 6) → 'project'
      4. None (can't determine)
    """
    t = title.lower()
    if re.search(r'\blp\b', t):
        return 'lp'
    if 'project' in t:
        return 'project'
    if col <= 3:
        return 'lp'
    if col >= 5:
        return 'project'
    return None


_CASCADIA_RE = re.compile(r'cascadia', re.IGNORECASE)


def _is_cascadia_mislabel(title: str, sheet_name: str) -> bool:
    """
    Detect the Cascadia copy-paste defect on Verde Vista / Serene Springs sheets.

    A scroll block is considered mislabelled when its title mentions 'cascadia'
    but the sheet is not a Cascadia sheet.
    """
    if not _CASCADIA_RE.search(title):
        return False
    sheet_lower = sheet_name.lower()
    if 'cascadia' in sheet_lower:
        return False  # Legitimate Cascadia sheet
    return True


def parse_excel(file_or_path) -> dict:
    """
    Parse the Excel workbook and return the v2 data structure.

    Returns:
        {
            sheet_name: {
                'pages': {
                    'lp':      { ... } | None,
                    'project': { ... } | None,
                },
                'flags': [str, ...]
            },
            ...
        }
    """
    wb = openpyxl.load_workbook(file_or_path, data_only=True, read_only=True)
    log.info('opened workbook: %d sheets', len(wb.sheetnames))
    result = {}

    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in _SKIP_SHEETS:
            log.debug('skipping sheet %r', sheet_name)
            continue

        ws = wb[sheet_name]
        flags = []

        blocks = _discover_blocks(ws)
        if not blocks:
            log.debug('[%s] no blocks found — skipping', sheet_name)
            continue

        log.info('[%s] found %d block(s): %s', sheet_name, len(blocks),
                 [(b['kind'], f"r{b['header_row']}c{b['key_col']}") for b in blocks])

        # ── Parse all blocks ──────────────────────────────────────────────────
        click_id_blocks: list[dict] = []
        event_name_blocks: list[dict] = []

        for blk in blocks:
            if blk['kind'] == 'click_id':
                parsed = _parse_click_id_block(ws, blk['header_row'], blk['key_col'])
                parsed['key_col'] = blk['key_col']
                log.debug('[%s] click_id block title=%r  items=%d  total_clicks=%d',
                          sheet_name, parsed.get('title'), len(parsed.get('items', [])),
                          parsed.get('total_clicks', 0))
                click_id_blocks.append(parsed)
            else:
                parsed = _parse_event_name_block(ws, blk['header_row'], blk['key_col'])
                parsed['key_col'] = blk['key_col']

                # Detect and discard the Cascadia mislabel defect
                if _is_cascadia_mislabel(parsed['title'], sheet_name):
                    flags.append(
                        f"Scroll block '{parsed['title']}' looks like a Cascadia copy-paste "
                        f"on sheet '{sheet_name}' — discarded."
                    )
                    log.warning('[%s] Cascadia mislabel detected: block title=%r — discarded',
                                sheet_name, parsed['title'])
                    continue

                log.debug('[%s] event_name block title=%r  pv_users=%d  buckets=%s',
                          sheet_name, parsed.get('title'),
                          parsed.get('page_view_users', 0),
                          [f"{int(d*100)}%={u}" for d, u in parsed.get('buckets', [])])
                event_name_blocks.append(parsed)

        if not click_id_blocks and not event_name_blocks:
            continue

        # ── Map blocks to page types ──────────────────────────────────────────
        pages: dict[str, dict] = {}

        for blk in click_id_blocks:
            title = blk['title'] or f'{sheet_name} LP'
            pt = _assign_page_type(title, blk['key_col'])
            if pt is None:
                # Fall back: first block → lp, second → project
                pt = 'lp' if not pages else 'project'

            # Sort items by clicks desc (export row order is unreliable)
            items = sorted(blk['items'], key=lambda x: x['clicks'], reverse=True)

            pages[pt] = {
                'title': title,
                'total_clicks': blk['total_clicks'],
                'total_users': blk['total_users'],
                'items': items,
                'scroll': None,
                'mode': 'clicks_only',
            }

        # ── Attach scroll data ────────────────────────────────────────────────
        for blk in event_name_blocks:
            if not blk['buckets']:
                continue

            title = blk['title']
            pt = _assign_page_type(title, blk['key_col'])

            # Try both page types if we can't determine from title
            targets = [pt] if pt is not None else list(pages.keys())

            attached = False
            for candidate in targets:
                if candidate in pages:
                    pages[candidate]['scroll'] = {
                        'page_view_users': blk['page_view_users'],
                        'buckets': blk['buckets'],
                    }
                    pages[candidate]['mode'] = 'full'
                    attached = True
                    break

            if not attached:
                # Scroll-only block (no matching click_id block on this sheet)
                # Still store the page for portfolio attention metrics
                if pt not in pages:
                    pages[pt or 'lp'] = {
                        'title': title,
                        'total_clicks': 0,
                        'total_users': 0,
                        'items': [],
                        'scroll': {
                            'page_view_users': blk['page_view_users'],
                            'buckets': blk['buckets'],
                        },
                        'mode': 'clicks_only',
                    }

        # ── Check for non-monotonic scroll curves ─────────────────────────────
        for pt, page in pages.items():
            scroll = page.get('scroll')
            if scroll and scroll['buckets']:
                prev_users = scroll['page_view_users']
                non_mono = False
                for depth, users in scroll['buckets']:
                    if users > prev_users:
                        flags.append(
                            f"Non-monotonic scroll curve on {sheet_name}/{pt}: "
                            f"{depth:.0%} bucket has {users:,} users > previous {prev_users:,}."
                        )
                        non_mono = True
                        break
                    prev_users = users
                if non_mono:
                    log.warning('[%s/%s] non-monotonic scroll curve — suppressing attention model',
                                sheet_name, pt)
                    page['scroll'] = None
                    page['mode'] = 'clicks_only'

        if pages:
            result[sheet_name] = {'pages': pages, 'flags': flags}
            for pt, page in pages.items():
                log.info('[%s/%s] mode=%s  items=%d  pv_users=%s  buckets=%d  flags=%d',
                         sheet_name, pt, page['mode'], len(page['items']),
                         page['scroll']['page_view_users'] if page['scroll'] else 'n/a',
                         len(page['scroll']['buckets']) if page['scroll'] else 0,
                         len(flags))
        else:
            log.debug('[%s] no usable pages after parsing', sheet_name)

    log.info('parse_excel complete: %d sheets parsed', len(result))
    wb.close()
    return result


def get_sheet_names(file_or_path) -> list:
    """Return list of non-index sheet names from the Excel workbook."""
    wb = openpyxl.load_workbook(file_or_path, data_only=True, read_only=True)
    names = [s for s in wb.sheetnames if s.strip().lower() not in _SKIP_SHEETS]
    wb.close()
    return names
